import re
import logging
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except:
    raise EnvironmentError("当前python环境尚未安装openpyxl，请先执行python -m pip install openpyxl")


logger = logging.getLogger()


def list_completed_case_files(input_path: Path) -> list[Path]:
    """Read the canonical case file list from _progress.md."""
    progress_path = input_path / "_progress.md"
    if not progress_path.is_file():
        raise ValueError(f"缺少用例文件清单: {progress_path}")

    rows = []
    for raw_line in progress_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not (line.startswith("|") and line.endswith("|")):
            continue
        columns = [column.strip() for column in line.strip("|").split("|")]
        if len(columns) < 4 or columns[0] in {"序号", "---"}:
            continue
        if all(set(column) <= {"-", ":", " "} for column in columns):
            continue
        rows.append(columns)

    completed_files = []
    seen = set()
    root = input_path.resolve()
    for columns in rows:
        relative_name = columns[2].strip().strip('`')
        status = columns[3].strip().strip("`")
        if status != "已完成":
            continue
        case_path = (input_path / relative_name).resolve()
        if case_path.parent != root and root not in case_path.parents:
            raise ValueError(f"用例清单包含越界路径: {relative_name}")
        if case_path.suffix.lower() != ".md":
            raise ValueError(f"用例清单仅支持 Markdown 文件: {relative_name}")
        if case_path in seen:
            raise ValueError(f"用例清单包含重复文件: {relative_name}")
        if not case_path.is_file():
            raise ValueError(f"用例清单中的文件不存在: {relative_name}")
        seen.add(case_path)
        completed_files.append(case_path)

    if not completed_files:
        raise ValueError(f"用例清单中没有状态为“已完成”的文件: {progress_path}")
    return completed_files


def _replace_br(text):
    text = text.replace(r'<br>', ' ')
    return re.sub(r"\n\s*", "\n", text, re.DOTALL)


def parse_test_cases_from_markdown(md_text: str, start_index: int = 1) -> dict:
    """从 Markdown 文本中解析测试用例和知识信息

    Args:
        md_text: Markdown 文本内容。
        start_index: 自动生成 ID 时的起始序号，用于跨文件全局递增。
    """
    knowledge = ""
    test_cases = []

    # 提取知识库信息
    # 匹配 # 知识库信息 之后直到下一个大标题或分割线的内容
    knowledge_match = re.search(r'# 测试概述\s*\n(.*?)(?=\n---|\n##|\n# |$)', md_text, re.DOTALL)
    if knowledge_match:
        knowledge = knowledge_match.group(1).strip()

    # 类型映射
    type_mapping = {
        "正常": "functional",
        "功能": "functional",
        "边界": "boundary",
        "异常": "error",
        "安全": "security",
        "functional": "functional",
        "boundary": "boundary",
        "error": "error",
        "security": "security"
    }

    # 每条用例使用标题、用例类型、优先级三个独立字段。
    segments = re.split(r'\n(?=#{2,}\s)', md_text)

    for seg in segments:
        title_match = re.match(r'#{2,}\s+([^\n]+)', seg)
        if not title_match or not re.search(r'\*\*前置条件\*\*[:：]', seg):
            continue
        tc_name = re.sub(r'\*\*', '', title_match.group(1)).strip(" #:")

        def find_id(text):
            match = re.search(r'[A-Za-z_]+(?:[_-][A-Za-z_]+)*[_-]\d+', text)
            if match:
                return match.group(0)
            return None

        # --- 用例编号提取（三级回退，向后兼容）---
        # Tier 1: 新格式 — 从 **用例编号** 字段提取
        case_id_match = re.search(r'\*\*用例编号\*\*[:：][^\S\n]*([^\n]+)', seg)
        if case_id_match and case_id_match.group(1).strip():
            tc_id = case_id_match.group(1).strip()
        else:
            # Tier 2: 旧格式 — 从标题中提取编号，并清理标题
            tc_id = find_id(tc_name)
            if tc_id:
                tc_name = re.sub(r'\s*' + re.escape(tc_id) + r'\s*', '', tc_name).strip()
            else:
                # Tier 3: 兜底 — 自动生成
                tc_id = f"TC_{(start_index + len(test_cases)):0>3}"

        tc = {
            "id": tc_id,
            "name": tc_name,
            "priority": "P1",
            "type": "functional",
            "precondition": "",
            "steps": "",
            "expected": "",
            "test_data": "",
            "remark": "",
            "case_knowledge": knowledge
        }

        # 提取各个字段
        def get_field(pattern, text, default=None, multiline=False):
            if multiline:
                m = re.search(pattern, text, re.MULTILINE | re.DOTALL)
            else:
                m = re.search(pattern, text)
            return _replace_br(m.group(1).strip("*- ")) if m else default

        body_priority = get_field(r'\*\*优先级\*\*[:：]\s*([^\n]*)', seg)
        if body_priority:
            tc["priority"] = body_priority.upper()

        body_type = get_field(r'\*\*用例类型\*\*[:：]\s*([^\n]*)', seg)
        if body_type:
            # 同样对 body 中的类型进行映射
            body_type = body_type.replace("测试", "")
            body_type_lower = body_type.lower()
            if body_type_lower in type_mapping:
                tc["type"] = type_mapping[body_type_lower]
            else:
                for k, v in type_mapping.items():
                    if k in body_type_lower:
                        tc["type"] = v
                        break

        tc["precondition"] = get_field(r'\*\*前置条件\*\*[:：]\s*(.*?)(?=测试步骤)', seg, "", True)
        tc["test_data"] = get_field(r'\*\*测试数据\*\*[:：]\s*(.*?)(?=备注)', seg, "", True)
        tc["remark"] = get_field(r'\*\*备注\*\*[:：]\s*(.*?)(?=---|#|\Z)', seg, "", True)

        # 提取步骤 (多行)
        steps_match = re.search(r'\*\*测试步骤\*\*[:：]\s*(.*?)(?=\n\*\*|\n---|\n##|$)', seg, re.DOTALL)
        if steps_match:
            steps_content = steps_match.group(1).strip()

            # 检查是否包含 Markdown 表格结构
            if '|' in steps_content and '---' in steps_content:
                table_lines = [line.strip() for line in steps_content.split('\n') if line.strip()]
                steps_list = []
                expected_list = []

                for line in table_lines:
                    if re.match(r'^[|\s:-]+$', line):
                        continue
                    if re.search(r'步骤|step|操作|operation|预期|expected', line, re.I):
                        if line == table_lines[0]:
                            continue

                    cols = [c.strip() for c in line.strip('|').split('|')]
                    if len(cols) >= 2:
                        step_num = cols[0]
                        action = cols[1]
                        expected = cols[2] if len(cols) >= 3 else ""

                        prefix = f"{step_num}. " if step_num and not step_num.endswith('.') else f"{step_num} "
                        if not step_num: prefix = ""

                        steps_list.append(f"{action}")
                        if expected:
                            expected_list.append(f"{expected}")

                tc["steps"] = _replace_br("\n".join(steps_list))
                tc["expected"] = _replace_br("\n".join(expected_list))
            else:
                # 不是表格模式
                steps = get_field(r'\*\*测试步骤\*\*[:：]\s*(.*?)(?=预期结果)', _replace_br(seg), "", True)
                if steps:
                    tc["steps"] = steps
                else:
                    tc["steps"] = _replace_br(steps_content)
                expected = get_field(r'\*\*预期结果\*\*[:：]\s*(.*?)(?=测试数据)', _replace_br(seg), "", True)
                if expected:
                    tc["expected"] = expected
                else:
                    expected_match = re.search(r'\*\*预期结果\*\*[:：]\s*(.*?)(?=\n\*\*|\n---|\n##|$)', seg, re.DOTALL)
                    if expected_match:
                        tc["expected"] = _replace_br(expected_match.group(1).strip())
            test_cases.append(tc)
        else:
            if len(test_cases) > 0:
                logger.warning(f"未找到测试步骤标题: {tc_name}")

    return {"test_cases": test_cases, "knowledge": knowledge}


def convert_markdown_cases_to_excel(input_dir: str, output_path: str) -> str:
    """将 _progress.md 清单中的已完成用例转换为 Excel 文件。

    Args:
        input_dir: 包含 Markdown 用例文件的目录路径。
        output_path: 输出的 Excel 文件路径（.xlsx）。

    Returns:
        生成的 Excel 文件路径。
    """
    # Use one absolute base path throughout. Case files are resolved while
    # reading _progress.md, so a relative input path previously broke
    # md_file.relative_to(input_path) during module-name construction.
    input_path = Path(input_dir).resolve()
    if not input_path.is_dir():
        raise ValueError(f"输入路径不是有效的目录: {input_dir}")

    md_files = list_completed_case_files(input_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    TYPE_LABELS = {
        "functional": "功能",
        "boundary": "边界",
        "error": "异常",
        "security": "安全",
    }

    headers = ["ID", "模块", "优先级", "类型", "标题", "前置条件", "步骤", "预期", "测试数据", "备注"]
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="D9E2F3")
    cell_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    row = 2
    global_index = 1
    module_ranges: dict[str, list[int]] = {}

    for md_file in md_files:
        if "progress.md" in md_file.name.lower() or "test-case-summary.md" in md_file.name.lower():
            continue
        relative = md_file.relative_to(input_path)
        parts = list(relative.parent.parts) + [relative.stem]
        module = "-".join(parts) if parts else relative.stem
        module = module.replace("测试用例", "").replace("用例", "").strip("-")

        try:
            content = md_file.read_text(encoding="utf-8")
            result = parse_test_cases_from_markdown(content, start_index=global_index)
            test_cases = result.get("test_cases", [])
            global_index += len(test_cases)
        except Exception as e:
            logger.warning(f"解析文件失败: {md_file}, 错误: {e}")
            continue

        for tc in test_cases:
            tc_type = TYPE_LABELS.get(tc.get("type", ""), tc.get("type", ""))
            ws.cell(row=row, column=1, value=tc.get("id", ""))
            ws.cell(row=row, column=2, value=module)
            ws.cell(row=row, column=3, value=tc.get("priority", ""))
            ws.cell(row=row, column=4, value=tc_type)
            ws.cell(row=row, column=5, value=tc.get("name", ""))
            ws.cell(row=row, column=6, value=tc.get("precondition", ""))
            ws.cell(row=row, column=7, value=tc.get("steps", ""))
            ws.cell(row=row, column=8, value=tc.get("expected", ""))
            ws.cell(row=row, column=9, value=tc.get("test_data", ""))
            ws.cell(row=row, column=10, value=tc.get("remark", ""))

            if module not in module_ranges:
                module_ranges[module] = [row, row]
            else:
                module_ranges[module][1] = row

            row += 1

    # 合并相同模块的单元格
    merge_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for start_row, end_row in module_ranges.values():
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.cell(row=start_row, column=2).alignment = merge_center

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
            [14, 28, 10, 10, 34, 40, 40, 42, 28, 36][col_idx - 1]

    for r in range(2, row):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = wrap_alignment
            ws.cell(row=r, column=c).border = cell_border
        ws.row_dimensions[r].height = 90

    last_col = chr(64 + len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{row - 1}"
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Excel 文件已保存: {output_path}，共 {row - 2} 条用例")
    return output_path


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="将 _progress.md 清单中的已完成 Markdown 测试用例转换为 Excel 文件"
    )
    parser.add_argument(
        "input_dir",
        help="包含 _progress.md 和 Markdown 用例文件的目录路径",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="输出的 Excel 文件路径（默认在输入目录下生成 testcases.xlsx）",
    )
    args = parser.parse_args()

    input_dir = args.input_dir
    output_path = args.output or str(Path(input_dir) / "testcases.xlsx")

    try:
        result = convert_markdown_cases_to_excel(input_dir, output_path)
        print(f"转换完成! output path: {result}")
    except Exception as e:
        logger.error(f"转换失败: {e}")
        raise


if __name__ == "__main__":
    main()
